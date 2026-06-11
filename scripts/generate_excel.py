"""
generate_excel.py

Generates dummy_profiles.xlsx with all dummy profile data.
Output: /Users/sc.park/Desktop/3_KAIST/2026-Spring/1_데이터마이닝/4_TeamProject/dummy_profiles.xlsx

Usage:
    python3 scripts/generate_excel.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def drive_url(file_id):
    return f"https://drive.google.com/thumbnail?id={file_id}&sz=w800"

# All dummy profiles
PROFILES = [
    # (userId, gender, nickname, age, archetypeId, archetypeName, archetypeCategory, styleTemp, keywords, imageUrls)
    ("dummy_1_m", "M", "준혁", 25, 1, "미니멀 시티보이", "미니멀", 78,
     "미니멀,모던캐주얼,깔끔,시티무드", []),

    ("dummy_2_f", "F", "지윤", 24, 2, "미니멀 시크걸", "미니멀", 76,
     "미니멀,시크,단정함,모노톤", []),

    ("dummy_3_m", "M", "민재", 23, 3, "클린 베이직", "미니멀", 58,
     "베이직,깔끔,무난,데일리",
     [drive_url("14tgKjWZShIu3k5YcCBnlUF8wo9819I0L")]),

    ("dummy_3_f", "F", "수빈", 22, 3, "클린 베이직", "미니멀", 55,
     "베이직,깔끔,무난,데일리",
     [drive_url("1Z47Q-lcJ_QB2gIjysbs5YY41lvzOJ5Dy"),
      drive_url("1zOLGyPlUYiZQt_jrPzSeAFS8qD1juNH-"),
      drive_url("1xt-goBh_9JG2m3okZ8QA-yn-z4KO4meN")]),

    ("dummy_4_m", "M", "서진", 26, 4, "모노톤 무드", "미니멀", 72,
     "블랙&화이트,무채색,절제,심플",
     [drive_url("1usR832-MMx2kcwF6Xy6EL-w0umCaD06h")]),

    ("dummy_4_f", "F", "하은", 25, 4, "모노톤 무드", "미니멀", 68,
     "블랙&화이트,무채색,절제,심플",
     [drive_url("1DEMVTK2z00seIfCMSvuP5VCwnWlpNrpH")]),

    ("dummy_5_m", "M", "현우", 27, 5, "쿨톤 미니멀", "미니멀", 74,
     "블루그레이,차가운톤,슬림핏,절제미",
     [drive_url("1n2YweCbzQvn2fDABNJguQeJRaEQ0NbjD")]),

    ("dummy_5_f", "F", "예린", 24, 5, "쿨톤 미니멀", "미니멀", 76,
     "블루그레이,차가운톤,슬림핏,절제미",
     [drive_url("1ac-64jGh7lsppRBSP7px3QxUN6hqybez")]),

    ("dummy_6_m", "M", "도윤", 24, 6, "릴렉스 캐주얼", "캐주얼", 47,
     "캐주얼,편안함,자연스러움,데일리",
     [drive_url("12JxpXADbCQ9a36vH5clAdvd8cwZ9xU6M")]),

    ("dummy_6_f", "F", "다은", 23, 6, "릴렉스 캐주얼", "캐주얼", 44,
     "캐주얼,편안함,자연스러움,데일리",
     [drive_url("1E3Eg1n-PF52jeYz7ok5UkhhdJbjm4XJ_")]),

    ("dummy_7_m", "M", "지호", 22, 7, "스포티 프레시", "캐주얼", 53,
     "스포츠믹스,활동적,컬러풀,에너지",
     [drive_url("1gGWTyMoJEqdaEvyOswjmy9jFD1jeW9ve")]),

    ("dummy_7_f", "F", "유진", 23, 7, "스포티 프레시", "캐주얼", 50,
     "스포츠믹스,활동적,컬러풀,에너지",
     [drive_url("1eKLX-5V-wLr05JBrJWVmcVdEjh4gZIIs")]),

    ("dummy_8_m", "M", "태양", 25, 8, "코지 무드", "캐주얼", 48,
     "니트,따뜻함,라운지,오버핏",
     [drive_url("1dgdSlB9rX7j4XceFROC2ISnaOsLFw3Yg")]),

    ("dummy_8_f", "F", "소희", 24, 8, "코지 무드", "캐주얼", 46,
     "니트,따뜻함,라운지,오버핏",
     [drive_url("1-uWPnSSODSKGFI0mh1LdiZUWUMyljvAi")]),

    ("dummy_9_m", "M", "우진", 23, 9, "서프 캐주얼", "캐주얼", 42,
     "서핑,리조트,여유로움,썸머바이브",
     [drive_url("1zq7gDd-PaZAr7bZmUqsHkWVsrf9xLItY")]),

    ("dummy_9_f", "F", "민지", 22, 9, "서프 캐주얼", "캐주얼", 39,
     "서핑,리조트,여유로움,썸머바이브",
     [drive_url("1stl3zTIM7A3csQfpz_4T9COaLGSQaC5H")]),

    ("dummy_10_m", "M", "성민", 22, 10, "캠퍼스 프레피", "캐주얼", 62,
     "캠퍼스,프레피,니트조끼,체크",
     [drive_url("1YHfjCJKRoSQgKwrX0TxsPVxv29vZGQ-v")]),

    ("dummy_10_f", "F", "채원", 23, 10, "캠퍼스 프레피", "캐주얼", 60,
     "캠퍼스,프레피,니트조끼,체크",
     [drive_url("1deeZUW9wEfOFEmYu0q2XVZ8BHgQi4eg-")]),

    ("dummy_11_m", "M", "재현", 24, 11, "어반 스트릿", "스트릿", 68,
     "스트릿,오버핏,그래픽,하이탑",
     [drive_url("1fztVfLqo-F0o-QLTdXpEV70-cNrWhTJI")]),

    ("dummy_11_f", "F", "하윤", 25, 11, "어반 스트릿", "스트릿", 65,
     "스트릿,오버핏,그래픽,하이탑",
     [drive_url("10deAN-deGj2GWspcIkkjo_XvMWAu2pcL")]),

    ("dummy_13_m", "M", "건우", 26, 13, "테크웨어 무드", "스트릿", 82,
     "테크웨어,기능성,올블랙,미래적",
     [drive_url("1_XmGnfvvh6NHN9zJ8o92Fcs3_tX11zHJ")]),

    ("dummy_13_f", "F", "서연", 24, 13, "테크웨어 무드", "스트릿", 79,
     "테크웨어,기능성,올블랙,미래적",
     [drive_url("1VKmA19Ghu4_Nlkk84Q2YXWH6HaViYMgb")]),

    ("dummy_14_m", "M", "시우", 23, 14, "스케이터 바이브", "스트릿", 57,
     "스케이트,루즈핏,그런지,로고",
     [drive_url("1879G2D9RTR1pkKEMfpJj3Zd3GcEkpMaH")]),

    ("dummy_14_f", "F", "지수", 22, 14, "스케이터 바이브", "스트릿", 54,
     "스케이트,루즈핏,그런지,로고",
     [drive_url("179_XE9oHv9Jg4yaquAU7NeTftE7as-XH")]),

    ("dummy_15_m", "M", "은호", 27, 15, "모던 클래식", "클래식", 78,
     "클래식,단정함,셔츠,슬랙스",
     [drive_url("1sY7MVxeIk3oARdZfunXNA9PTeA-TIf2l")]),

    ("dummy_15_f", "F", "서현", 26, 15, "모던 클래식", "클래식", 75,
     "클래식,단정함,셔츠,슬랙스",
     [drive_url("11_5hr7JxTYU3UjS7gjOfnnbTrmEp_D-j")]),

    ("dummy_16_m", "M", "정우", 25, 16, "소프트 클래식", "클래식", 72,
     "클래식,부드러움,뉴트럴,여성스러움",
     [drive_url("1_3E6rAtxHCT4UTyFAjY9IpUqqxPss1Oe")]),

    ("dummy_17_m", "M", "승현", 28, 17, "댄디 포멀", "클래식", 91,
     "포멀,수트,정장감,세련됨",
     [drive_url("1Y2hXWhqxtV9UBkwHqvmFba5p90GB4Puv"),
      drive_url("15WJjbW31VWZm9pRss_QaSwcu30C9nB7u")]),

    ("dummy_18_m", "M", "민서", 27, 18, "비즈캐주얼 무드", "클래식", 73,
     "비즈캐주얼,오피스룩,깔끔,신뢰감",
     [drive_url("1hNbck0Hzo2DZ2iJ7gsEbMInqaHpesZZb")]),

    ("dummy_19_m", "M", "예준", 26, 19, "다크 아티스틱", "아티스틱", 90,
     "블랙,아방가르드,질감,실루엣",
     [drive_url("16uWS6uULQSYQYVgga2JK4fB4CkLavI9Q")]),

    ("dummy_19_f", "F", "나윤", 25, 19, "다크 아티스틱", "아티스틱", 88,
     "블랙,아방가르드,질감,실루엣",
     [drive_url("1L1dwhnJ_IBrsnc90XV79gKrGVw6ICmkW")]),

    ("dummy_20_m", "M", "유찬", 24, 20, "감성 내추럴", "아티스틱", 56,
     "어스톤,내추럴,오가닉,차분함",
     [drive_url("1AodTw6Pyr9KLWpJHBpdxsjFmnDkJ0Df8")]),

    ("dummy_20_f", "F", "혜진", 23, 20, "감성 내추럴", "아티스틱", 58,
     "어스톤,내추럴,오가닉,차분함",
     [drive_url("1SWnMLSlLMNFxczfh2ZIOJwz3kjsMyG5d")]),

    ("dummy_21_f", "F", "가은", 24, 21, "빈티지 로맨틱", "아티스틱", 77,
     "빈티지,레트로,플로럴,레이스",
     [drive_url("1WaKS81JPiM6z60k9k_kzac1J9yo7o2wn"),
      drive_url("1sxWJIlTwGDk8veyEJyN6l7ERZu2teZCx")]),

    ("dummy_22_m", "M", "동현", 25, 22, "뉴트로 무드", "아티스틱", 67,
     "레트로,컬러,올드스쿨,믹스매치",
     [drive_url("13gW72zMFFU2vOD2WJ8GxaumpBzmCOmLD")]),

    ("dummy_22_f", "F", "은서", 24, 22, "뉴트로 무드", "아티스틱", 64,
     "레트로,컬러,올드스쿨,믹스매치",
     [drive_url("1DQjUP_V5rqwS4bc-TuFfVPyeUn3v_Yrc")]),

    ("dummy_24_f", "F", "보라", 23, 24, "러블리 페미닌", "페미닌", 80,
     "페미닌,핑크,프릴,원피스", []),

    ("dummy_25_f", "F", "수아", 26, 25, "걸크러시 시크", "페미닌", 83,
     "걸크러시,파워풀,모던,숏컷", []),

    ("dummy_26_f", "F", "연서", 25, 26, "프렌치 시크", "페미닌", 72,
     "프렌치,에포트리스,베레모,스트라이프",
     [drive_url("1_CSRHeZv2rrV7xHUi9RhUG2Vz_mPRGE9"),
      drive_url("1qjaSCYNJNW7yb6eg56tn3M4nS1JqKW_K")]),

    ("dummy_27_f", "F", "다현", 27, 27, "글램 맥시멀", "맥시멀", 95,
     "화려함,악세서리,패턴,볼드컬러",
     [drive_url("1amdlpsSrgY-XwhkC0yqmv4TqGtKxt5UA"),
      drive_url("1QdKJsfTv9fTKsVLB_xPzjZQZe9rJKouI")]),

    ("dummy_29_m", "M", "하준", 26, 29, "아메카지 무드", "맥시멀", 68,
     "워크웨어,빈티지,내추럴,레이어드",
     [drive_url("1h4jyrxW-GquDkchgPqbXFK2KkVjc2HmX"),
      drive_url("1vBgrSQcmzS1BaNPrlNdFsGb9LCoWFBcn")]),

    ("dummy_30_m", "M", "준서", 23, 30, "컬러팝 무드", "맥시멀", 78,
     "비비드,컬러블록,대담,펀",
     [drive_url("1sG9iU16cZh6HcNiRj7J9Ju52XOuUPXQ7")]),

    ("dummy_30_f", "F", "하린", 22, 30, "컬러팝 무드", "맥시멀", 75,
     "비비드,컬러블록,대담,펀",
     [drive_url("1eBNuaNnbY3R22Azu6jX6aPDzwzEMDoAg")]),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "더미 프로필"

    # Headers
    headers = [
        "더미ID", "성별", "이름", "나이",
        "아키타입ID", "아키타입", "카테고리", "스타일온도",
        "키워드", "이미지URL1", "이미지URL2", "이미지URL3"
    ]

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows
    data_font = Font(name="맑은 고딕", size=10)
    male_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    female_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    for row_idx, p in enumerate(PROFILES, 2):
        user_id, gender, nickname, age, arch_id, arch_name, arch_cat, style_temp, keywords, urls = p
        row_data = [
            user_id, gender, nickname, age,
            arch_id, arch_name, arch_cat, style_temp,
            keywords,
            urls[0] if len(urls) > 0 else "",
            urls[1] if len(urls) > 1 else "",
            urls[2] if len(urls) > 2 else "",
        ]
        fill = male_fill if gender == "M" else female_fill
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill
            if col_idx in (1, 2, 4, 5, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    # Column widths
    col_widths = {
        1: 18,   # 더미ID
        2: 6,    # 성별
        3: 8,    # 이름
        4: 6,    # 나이
        5: 10,   # 아키타입ID
        6: 18,   # 아키타입
        7: 10,   # 카테고리
        8: 10,   # 스타일온도
        9: 35,   # 키워드
        10: 55,  # 이미지URL1
        11: 55,  # 이미지URL2
        12: 55,  # 이미지URL3
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(PROFILES) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="총 프로필 수").font = Font(bold=True, name="맑은 고딕")
    ws2.cell(row=1, column=2, value=len(PROFILES))

    male_count = sum(1 for p in PROFILES if p[1] == "M")
    female_count = sum(1 for p in PROFILES if p[1] == "F")
    ws2.cell(row=2, column=1, value="남성 프로필").font = Font(bold=True, name="맑은 고딕")
    ws2.cell(row=2, column=2, value=male_count)
    ws2.cell(row=3, column=1, value="여성 프로필").font = Font(bold=True, name="맑은 고딕")
    ws2.cell(row=3, column=2, value=female_count)

    # Count by category
    categories = {}
    for p in PROFILES:
        cat = p[6]
        categories[cat] = categories.get(cat, 0) + 1

    ws2.cell(row=5, column=1, value="카테고리별 분포").font = Font(bold=True, name="맑은 고딕", size=12)
    for i, (cat, count) in enumerate(sorted(categories.items()), 6):
        ws2.cell(row=i, column=1, value=cat).font = Font(name="맑은 고딕")
        ws2.cell(row=i, column=2, value=count)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10

    # Save
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "..",
        "dummy_profiles.xlsx"
    )
    output_path = os.path.normpath(output_path)
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    print(f"Total profiles: {len(PROFILES)} (M={male_count}, F={female_count})")


if __name__ == "__main__":
    main()
